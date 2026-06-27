"""
EDI 835 Remittance Processing Pipeline — Entry Point

Usage:
    uv run python main.py                    # Full pipeline (download + parse + report)
    uv run python main.py --download-only    # Only download from SFTP
    uv run python main.py --parse-only       # Parse already-downloaded files
    uv run python main.py --output-dir DIR   # Override output directory
    uv run python main.py --scheduler        # Run continuously every hour
"""

import argparse
import logging
import sys
import time
from pathlib import Path

from remit_pipeline.config import Config
from remit_pipeline.state_db import StateDB


def setup_logging() -> None:
    """Configure logging with timestamped, leveled output."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def run_pipeline_step(parse_only: bool, download_only: bool, output_dir: Path) -> int:
    """
    Run a single pipeline execution step:
    1. SFTP Check & Download (if not parse_only)
    2. Incremental Parse (if not download_only)
    3. Master Report Rebuild (if not download_only)

    Returns:
        Number of new files processed successfully, or -1 on error.
    """
    from remit_pipeline.edi_parser import parse_edi_835
    from remit_pipeline.report_builder import rebuild_report_from_cached_csvs
    from remit_pipeline.sftp_client import SFTPClient
    import pandas as pd

    logger = logging.getLogger("pipeline.step")
    db_path = Config.RAW_DATA_DIR.parent / "pipeline_state.db"
    
    new_files_count = 0
    files_to_parse = []

    with StateDB(db_path) as db:
        # Get list of already successfully processed files
        processed_set = db.get_processed_files()

        # Step 1: SFTP connection and download
        if not parse_only:
            try:
                with SFTPClient() as sftp:
                    remote_files = sftp.list_rmt_files_with_attrs()
                    
                    # Filter out already processed files (by name and size)
                    new_remote_files = [
                        (fname, size) for fname, size in remote_files
                        if (fname, size) not in processed_set
                    ]

                    if not new_remote_files:
                        logger.info("No new files found on SFTP server.")
                    else:
                        logger.info("Found %d new/modified files to download.", len(new_remote_files))
                        
                        # Download each file
                        for fname, size in new_remote_files:
                            try:
                                sftp.download_files([fname], overwrite=True)
                                files_to_parse.append((fname, size))
                            except Exception as e:
                                logger.error("Failed to download '%s': %s", fname, e)
            except Exception as e:
                logger.error("SFTP operations failed: %s", e)
                if not download_only:
                    logger.info("Proceeding with already downloaded local files...")
        else:
            # Parse only mode: scan local RAW_DATA_DIR
            logger.info("Scanning local raw data directory: %s", Config.RAW_DATA_DIR)
            local_files = sorted([
                f for f in Config.RAW_DATA_DIR.iterdir()
                if f.is_file() and f.name.lower().endswith(".rmt")
            ])
            for f in local_files:
                size = f.stat().st_size
                if (f.name, size) not in processed_set:
                    files_to_parse.append((f.name, size))

        if download_only:
            return len(files_to_parse)

        # Step 2: Incremental Parse
        if files_to_parse:
            logger.info("Parsing %d new/modified files...", len(files_to_parse))
            for fname, size in files_to_parse:
                filepath = Config.RAW_DATA_DIR / fname
                try:
                    records = parse_edi_835(str(filepath), fname)
                    if records:
                        # Save per-file CSV
                        csv_name = Path(fname).stem + ".csv"
                        csv_path = Config.PROCESSED_DATA_DIR / csv_name
                        df_file = pd.DataFrame(records)
                        df_file.to_csv(csv_path, index=False)
                        logger.info("Saved CSV: %s (%d rows)", csv_path.name, len(records))
                        
                    # Mark successful in DB
                    db.mark_file(fname, size, "SUCCESS")
                    new_files_count += 1
                except Exception as e:
                    logger.error("Failed to parse '%s': %s", fname, e)
                    db.mark_file(fname, size, "FAILED")
        else:
            logger.info("No new files to parse.")

        # Step 3: Rebuild Master Report (if new files parsed OR if master report is missing)
        # Find latest master report path
        today_str = time.strftime("%Y%m%d")
        report_path = Config.OUTPUT_DIR / f"Master_Remittance_Report_{today_str}.xlsx"
        
        if new_files_count > 0 or not report_path.exists():
            logger.info("Generating/rebuilding Master Excel report...")
            try:
                output_path = rebuild_report_from_cached_csvs(
                    processed_dir=Config.PROCESSED_DATA_DIR,
                    output_dir=output_dir,
                )
                if output_path:
                    _print_summary(output_path)
            except Exception as e:
                logger.error("Failed to rebuild Master Excel report: %s", e)
        else:
            logger.info("Master report is up to date. Rebuilding skipped.")

    return new_files_count


def _print_summary(output_path: str) -> None:
    """Print a final summary of the generated report."""
    from openpyxl import load_workbook

    path = Path(output_path)
    size = path.stat().st_size

    wb = load_workbook(str(path), read_only=True)
    sheet_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row - 4  # Subtract header rows
        sheet_info.append((name, rows))
    wb.close()

    print()
    print("=" * 64)
    print("  ✓ Master Report Generated")
    print("=" * 64)
    print(f"  Output file: {output_path}")
    print(f"  File size:   {size:,} bytes ({size / 1024:.1f} KB)")
    print()
    for name, rows in sheet_info:
        print(f"  Sheet: {name:30s} → {rows:,} rows")
    print("=" * 64)


def main():
    parser = argparse.ArgumentParser(
        description="EDI 835 Remittance Processing Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  uv run python main.py                  Full pipeline (download + parse + report)
  uv run python main.py --download-only  Only download .rmt files from SFTP
  uv run python main.py --parse-only     Parse local files and generate Excel report
  uv run python main.py --output-dir DIR Override output directory
  uv run python main.py --scheduler      Run scheduler loop every 60 minutes
        """,
    )
    parser.add_argument(
        "--download-only",
        action="store_true",
        help="Only connect to SFTP and download .rmt files (Steps 2–3)",
    )
    parser.add_argument(
        "--parse-only",
        action="store_true",
        help="Skip SFTP download; parse already-downloaded files (Steps 4–5)",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Override the default output directory",
    )
    parser.add_argument(
        "--scheduler",
        action="store_true",
        help="Run in scheduler daemon mode, checking for new files every hour",
    )
    args = parser.parse_args()

    # Validate mutually exclusive flags
    if args.download_only and args.parse_only:
        print("ERROR: --download-only and --parse-only are mutually exclusive.")
        sys.exit(1)
    if args.scheduler and args.download_only:
        print("ERROR: --scheduler and --download-only are mutually exclusive.")
        sys.exit(1)

    # Setup logging
    setup_logging()
    logger = logging.getLogger("pipeline")

    if args.output_dir:
        Config.OUTPUT_DIR = Path(args.output_dir)

    Config.ensure_directories()

    print()
    print("=" * 64)
    print("  EDI 835 Remittance Processing Pipeline")
    print("=" * 64)
    print(f"  Raw data dir:       {Config.RAW_DATA_DIR}")
    print(f"  Processed data dir: {Config.PROCESSED_DATA_DIR}")
    print(f"  Output dir:         {Config.OUTPUT_DIR}")
    print(f"  Scheduler mode:     {'ACTIVE (every 60 min)' if args.scheduler else 'INACTIVE (one-shot)'}")
    print(f"  Mode:               {'Download Only' if args.download_only else 'Parse Only' if args.parse_only else 'Full Pipeline'}")
    print("=" * 64)
    print()

    # Run scheduler loop or single step
    if args.scheduler:
        logger.info("Initializing scheduler daemon loop...")
        while True:
            try:
                run_pipeline_step(
                    parse_only=args.parse_only,
                    download_only=False,
                    output_dir=Config.OUTPUT_DIR,
                )
            except Exception as e:
                logger.error("Error in scheduler execution step: %s", e)
            logger.info("Daemon sleep: waiting 60 minutes for next check...")
            time.sleep(3600)
    else:
        # Run one-shot
        logger.info("Executing single one-shot run...")
        count = run_pipeline_step(
            parse_only=args.parse_only,
            download_only=args.download_only,
            output_dir=Config.OUTPUT_DIR,
        )
        if count < 0:
            logger.error("Pipeline run failed.")
            sys.exit(1)
        logger.info("Pipeline run complete.")


if __name__ == "__main__":
    main()
