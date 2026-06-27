"""
Master Remittance Excel Report builder.

Produces a single Excel workbook with two sheets:
  1. "Claim Summary"       — one row per CLP (claim), matching the master template format
  2. "Service Line Detail" — one row per SVC (service date), full granularity

The claim summary is *derived* from the service-line detail via aggregation,
so the detail sheet is the single source of truth.
"""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from remit_pipeline.config import Config

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Claim status code → Transaction Type mapping
# ---------------------------------------------------------------------------
CLAIM_STATUS_MAP: Dict[str, str] = {
    "1": "Paid in Full",
    "2": "Adjusted",
    "3": "Denied",
    "4": "Denied/Adjusted",
    "19": "Paid in Full (Primary, Forwarded)",
    "20": "Adjusted (Secondary)",
    "21": "Denied (Secondary)",
    "22": "Reversal",
}

# ---------------------------------------------------------------------------
# Payer name → short Insurance label mapping
# ---------------------------------------------------------------------------
PAYER_INSURANCE_MAP: Dict[str, str] = {
    "UNITED": "United",
    "UNITEDHEALTHCARE": "United",
    "UHC": "United",
    "AETNA": "Aetna",
    "ANTHEM": "Anthem",
    "SENTARA": "Sentara",
    "HUMANA": "Humana",
    "CIGNA": "Cigna",
    "MOLINA": "Molina",
    "OPTIMA": "Optima",
    "MEDICAID": "Medicaid",
    "MEDICARE": "Medicare",
    "BCBS": "BCBS",
    "BLUE CROSS": "BCBS",
    "CAREFIRST": "CareFirst",
    "MAGELLAN": "Magellan",
    "PDN": "PDN",
    "TRICARE": "Tricare",
    "VIRGINIA PREMIER": "VA Premier",
}


def _derive_insurance(payer_name: str) -> str:
    """Map a payer name to a short insurance label."""
    if not payer_name:
        return ""
    upper = payer_name.upper()
    for keyword, label in PAYER_INSURANCE_MAP.items():
        if keyword in upper:
            return label
    return payer_name


def _derive_transaction_type(status_code) -> str:
    """Map claim status code to a human-readable transaction type."""
    return CLAIM_STATUS_MAP.get(str(status_code).strip(), "Other")


def _safe_float(val, default=0.0) -> float:
    """Convert a string to float, returning default on failure."""
    if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
        return default
    try:
        return float(str(val).replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return default


def _format_currency(val) -> str:
    """Format a numeric value as $#,##0.00 string."""
    return f"${_safe_float(val):,.2f}"


def _derive_month(first_dos: str) -> str:
    """Extract month prefix from a date string like '5/20/2026' → '5/'."""
    if not first_dos:
        return ""
    match = re.match(r"^(\d{1,2})/", str(first_dos))
    return match.group(1) + "/" if match else ""


def _base_procedure_code(proc_code: str) -> str:
    """Extract the base procedure code, stripping modifiers.
    'HC:T1019:76:UB' → 'T1019'
    'HC:T1005' → 'T1005'
    """
    if not proc_code:
        return ""
    parts = str(proc_code).split(":")
    # Find the Txxxx code (usually the second element after 'HC')
    for part in parts:
        if re.match(r"^[A-Z]\d{4}$", part):
            return part
    # Fallback: return everything after 'HC:'
    if parts[0] == "HC" and len(parts) > 1:
        return parts[1]
    return proc_code


# ============================================================================
# Build the Service Line Detail DataFrame (SVC-level — source of truth)
# ============================================================================

# Column order for the detail sheet
DETAIL_COLUMNS = [
    "Check/EFT Number",
    "Payment Date",
    "Payer Name",
    "Patient Last Name",
    "Patient First Name",
    "Member ID",
    "Claim Control Number (TCN)",
    "Transaction Type",
    "Claim Charge",
    "Claim Payment",
    "Claim Allowed",
    "First DOS",
    "Last DOS",
    "Service Date",
    "Procedure Code",
    "Units",
    "Billed Amount",
    "Paid Amount",
    "Line Allowed Amount (B6)",
    "Rendering Provider NPI",
    "Program / Waiver",
    "Service Line Ref (6R)",
    "Location Code (LU)",
    "CAS1_Group",
    "CAS1_Reason",
    "CAS1_Amount",
    "CAS2_Group",
    "CAS2_Reason",
    "CAS2_Amount",
    "CAS3_Group",
    "CAS3_Reason",
    "CAS3_Amount",
    "Remark Codes",
    "Source File",
]


def build_detail_dataframe(parsed_df: pd.DataFrame) -> pd.DataFrame:
    """
    Transform the raw parsed 835 DataFrame into a clean detail DataFrame.
    One row per SVC (service line / service date).
    Renames columns for clarity — distinguishes claim-level vs line-level amounts.
    """
    if parsed_df.empty:
        return pd.DataFrame(columns=DETAIL_COLUMNS)

    df = parsed_df.copy()

    # Add derived columns
    df["Insurance"] = df["Payer Name"].apply(lambda x: _derive_insurance(str(x)))
    df["Transaction Type"] = df["Claim Status Code"].apply(_derive_transaction_type)
    df["Procedure Code (Base)"] = df["Procedure Code"].apply(_base_procedure_code)

    # Rename to distinguish claim-level vs SVC-level amounts
    df = df.rename(columns={
        "Charge Amount": "Claim Charge",
        "Payment Amount": "Claim Payment",
        "Allowed Amount": "Claim Allowed",
        "Billed Amount (SVC)": "Billed Amount",
        "Paid Amount (SVC)": "Paid Amount",
        "Paid Amount per Line": "Line Allowed Amount (B6)",
    })

    # Reorder and select columns
    available = [c for c in DETAIL_COLUMNS if c in df.columns]
    df = df[available]

    logger.info("Detail DataFrame built: %d rows × %d columns", len(df), len(df.columns))
    return df


# ============================================================================
# Build the Claim Summary DataFrame (derived from detail)
# ============================================================================

# Column order for the claim summary sheet (matches master template + extensions)
SUMMARY_COLUMNS = [
    # --- Base 22 columns matching the existing master template ---
    "Batch",
    "Date",
    "Transaction",
    "Match Status",
    "Claim",
    "Transaction Type",
    "Charge",
    "Payment",
    "Allowed",
    "First Name",
    "Last Name",
    "First DOS",
    "Last DOS",
    "TCN",
    "Billed Hrs",
    "Paid Hrs",
    "Hrs Remaining",
    "Client",
    "Last Name, First Name",
    "Month",
    "Insurance",
    "Payment Value",
    # --- Additional 835 columns ---
    "Payer Name",
    "Member ID",
    "Rendering Provider NPI",
    "Program / Waiver",
    "Procedure Code",
    "Has Adjustment",
    "Total Adjustment",
    "Adjustment Codes",
    "Service Lines",
    "Source File",
]


def build_claim_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate the SVC-level detail into one row per claim (CLP).

    Groups by Claim Control Number (TCN) and aggregates:
    - Units, amounts: summed from SVC lines
    - Procedure codes: most common base code (or concatenated if mixed)
    - Adjustments: summarized with flag, total, and codes
    """
    if detail_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    rows: List[Dict] = []

    # Group by TCN (unique claim identifier)
    grouped = detail_df.groupby("Claim Control Number (TCN)", sort=False)

    for tcn, grp in grouped:
        first = grp.iloc[0]

        # --- Amounts (claim-level, same across all SVC rows) ---
        charge = _safe_float(first.get("Claim Charge", ""))
        payment = _safe_float(first.get("Claim Payment", ""))
        allowed = _safe_float(first.get("Claim Allowed", ""))

        # --- Units: sum across SVC lines ---
        total_units = grp["Units"].apply(_safe_float).sum()

        # --- Paid units: proportional ---
        billed_sum = grp["Billed Amount"].apply(_safe_float).sum()
        paid_sum = grp["Paid Amount"].apply(_safe_float).sum()
        if billed_sum > 0:
            paid_units = round(total_units * (paid_sum / billed_sum), 1)
        else:
            paid_units = 0.0

        # --- Procedure code: base code(s) ---
        base_codes = grp["Procedure Code"].apply(lambda x: _base_procedure_code(str(x))).dropna().unique()
        proc_code = ", ".join(sorted(set(base_codes))) if len(base_codes) > 0 else ""

        # --- Adjustments: aggregate CAS data ---
        cas_rows = grp[grp["CAS1_Group"].notna() & (grp["CAS1_Group"] != "")]
        has_adj = len(cas_rows) > 0
        total_adj = 0.0
        adj_codes_set = set()

        if has_adj:
            for _, cas_row in cas_rows.iterrows():
                for i in range(1, 4):  # CAS1, CAS2, CAS3
                    grp_val = str(cas_row.get(f"CAS{i}_Group", ""))
                    reason_val = str(cas_row.get(f"CAS{i}_Reason", ""))
                    amt_val = _safe_float(cas_row.get(f"CAS{i}_Amount", ""))
                    if grp_val and grp_val not in ("", "nan"):
                        # Clean up reason (remove trailing .0)
                        reason_clean = reason_val.replace(".0", "") if reason_val else ""
                        adj_codes_set.add(f"{grp_val}-{reason_clean}")
                        total_adj += amt_val

        adj_codes_str = "; ".join(sorted(adj_codes_set)) if adj_codes_set else ""

        # --- Patient info ---
        first_name = str(first.get("Patient First Name", ""))
        last_name = str(first.get("Patient Last Name", ""))
        first_dos = str(first.get("First DOS", ""))

        row = {
            "Batch": str(first.get("Check/EFT Number", "")),
            "Date": str(first.get("Payment Date", "")),
            "Transaction": "Payment",
            "Match Status": "Matched",
            "Claim": str(tcn),
            "Transaction Type": str(first.get("Transaction Type", "")),
            "Charge": charge,
            "Payment": payment,
            "Allowed": allowed,
            "First Name": first_name,
            "Last Name": last_name,
            "First DOS": first_dos,
            "Last DOS": str(first.get("Last DOS", "")),
            "TCN": str(tcn),
            "Billed Hrs": total_units,
            "Paid Hrs": paid_units,
            "Hrs Remaining": round(paid_units - total_units, 1),
            "Client": f"{last_name} {first_name}".strip(),
            "Last Name, First Name": f"{last_name}, {first_name}".strip(", "),
            "Month": _derive_month(first_dos),
            "Insurance": _derive_insurance(str(first.get("Payer Name", ""))),
            "Payment Value": payment,
            # --- Extended columns ---
            "Payer Name": str(first.get("Payer Name", "")),
            "Member ID": str(first.get("Member ID", "")),
            "Rendering Provider NPI": str(first.get("Rendering Provider NPI", "")),
            "Program / Waiver": str(first.get("Program / Waiver", "")),
            "Procedure Code": proc_code,
            "Has Adjustment": "Yes" if has_adj else "",
            "Total Adjustment": total_adj if has_adj else 0.0,
            "Adjustment Codes": adj_codes_str,
            "Service Lines": len(grp),
            "Source File": str(first.get("Source File", "")),
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=SUMMARY_COLUMNS)
    logger.info(
        "Claim Summary built: %d claims (from %d service lines)",
        len(df), len(detail_df),
    )
    return df


# ============================================================================
# Excel writer — two sheets in one workbook
# ============================================================================

# Style constants
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_DATA_FONT = Font(name="Calibri", size=11)
_SUMMARY_FONT = Font(name="Calibri", bold=True, size=12)
_SUMMARY_GREEN = Font(name="Calibri", bold=True, size=12, color="006100")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CURRENCY_FMT = '$#,##0.00'
_HRS_FMT = '0.0'
_INT_FMT = '0'


def _write_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    currency_cols: set,
    hrs_cols: set,
    int_cols: set,
    summary_row: bool = True,
    min_widths: Optional[Dict[str, float]] = None,
):
    """Write a DataFrame to a worksheet with full formatting."""
    ws = wb.create_sheet(title=sheet_name)
    columns = list(df.columns)
    num_cols = len(columns)
    min_widths = min_widths or {}

    data_start_row = 5 if summary_row else 2
    header_row = data_start_row - 1

    # --- Summary row (row 1) ---
    if summary_row:
        payment_col = None
        for col_name in ("Payment", "Paid Amount"):
            if col_name in df.columns:
                payment_col = col_name
                break
        total_payment = df[payment_col].apply(_safe_float).sum() if payment_col else 0

        date_col = None
        for col_name in ("Date", "Payment Date"):
            if col_name in df.columns:
                date_col = col_name
                break
        dates = df[date_col].dropna().unique() if date_col else []
        date_range = f"{min(dates)} – {max(dates)}" if len(dates) > 1 else (str(dates[0]) if len(dates) == 1 else "")

        ws.cell(row=1, column=1, value="Last uploaded remittance sheet is dated").font = _SUMMARY_FONT
        ws.cell(row=1, column=4, value=datetime.now().strftime("%m/%d/%Y")).font = Font(bold=True, size=12)
        ws.cell(row=1, column=5, value="Date Range:").font = _SUMMARY_FONT
        ws.cell(row=1, column=6, value=date_range).font = Font(bold=True, size=12)
        ws.cell(row=1, column=8, value="Total Payment:").font = _SUMMARY_FONT
        cell_total = ws.cell(row=1, column=9, value=total_payment)
        cell_total.font = _SUMMARY_GREEN
        cell_total.number_format = _CURRENCY_FMT

        ws.cell(row=2, column=1, value=f"Rows: {len(df):,}").font = Font(size=10, italic=True)
        ws.cell(row=2, column=4, value=f"Sheet: {sheet_name}").font = Font(size=10, italic=True)

    # --- Header row ---
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER

    # --- Build column type sets by index ---
    currency_idx = {i + 1 for i, c in enumerate(columns) if c in currency_cols}
    hrs_idx = {i + 1 for i, c in enumerate(columns) if c in hrs_cols}
    int_idx = {i + 1 for i, c in enumerate(columns) if c in int_cols}

    # --- Data rows ---
    for row_offset, (_, data_row) in enumerate(df.iterrows()):
        excel_row = data_start_row + row_offset
        for col_idx, col_name in enumerate(columns, 1):
            val = data_row[col_name]
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if col_idx in currency_idx:
                cell.value = _safe_float(val)
                cell.number_format = _CURRENCY_FMT
            elif col_idx in hrs_idx:
                cell.value = _safe_float(val)
                cell.number_format = _HRS_FMT
            elif col_idx in int_idx:
                cell.value = int(_safe_float(val))
                cell.number_format = _INT_FMT
            else:
                cell.value = str(val) if (val is not None and not (isinstance(val, float) and pd.isna(val))) else ""

    # --- Freeze panes ---
    ws.freeze_panes = ws.cell(row=data_start_row, column=2).coordinate

    # --- Auto-filter ---
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(df)}"

    # --- Column widths ---
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(str(col_name)) + 2

        for row_offset in range(min(100, len(df))):
            cell = ws.cell(row=data_start_row + row_offset, column=col_idx)
            if cell.value is not None:
                max_width = max(max_width, len(str(cell.value)) + 2)

        template_min = min_widths.get(col_name, 13)
        max_width = max(max_width, template_min)
        max_width = min(max_width, 40)
        ws.column_dimensions[col_letter].width = max_width


def write_master_excel(
    detail_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Write both sheets to a single Excel workbook.

    Sheet 1: "Claim Summary" — one row per claim (matches master template)
    Sheet 2: "Service Line Detail" — one row per SVC (full granularity)
    """
    output_dir = output_dir or Config.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    filename = "Master_Remittance_Report.xlsx"
    output_path = output_dir / filename

    wb = Workbook()
    # Remove the default sheet created by Workbook()
    wb.remove(wb.active)

    # --- Sheet 1: Claim Summary ---
    summary_currency = {"Charge", "Payment", "Allowed", "Payment Value", "Total Adjustment"}
    summary_hrs = {"Billed Hrs", "Paid Hrs", "Hrs Remaining"}
    summary_int = {"Service Lines"}
    summary_widths = {
        "Batch": 18, "Date": 12, "Transaction": 12, "Match Status": 13,
        "Claim": 16, "Transaction Type": 21, "Charge": 12, "Payment": 12,
        "Allowed": 12, "First Name": 16, "Last Name": 19, "First DOS": 11,
        "Last DOS": 11, "TCN": 16, "Billed Hrs": 11, "Paid Hrs": 10,
        "Hrs Remaining": 14, "Client": 19, "Last Name, First Name": 22,
        "Month": 8, "Insurance": 12, "Payment Value": 14,
        "Has Adjustment": 14, "Total Adjustment": 16, "Adjustment Codes": 18,
    }
    _write_sheet(
        wb, "Claim Summary", summary_df,
        currency_cols=summary_currency,
        hrs_cols=summary_hrs,
        int_cols=summary_int,
        summary_row=True,
        min_widths=summary_widths,
    )

    # --- Sheet 2: Service Line Detail ---
    detail_currency = {"Claim Charge", "Claim Payment", "Claim Allowed",
                       "Billed Amount", "Paid Amount", "Line Allowed Amount (B6)",
                       "CAS1_Amount", "CAS2_Amount", "CAS3_Amount"}
    detail_hrs = {"Units"}
    detail_int: set = set()
    detail_widths = {
        "Check/EFT Number": 18, "Claim Control Number (TCN)": 16,
        "Procedure Code": 18, "Line Allowed Amount (B6)": 24,
    }
    _write_sheet(
        wb, "Service Line Detail", detail_df,
        currency_cols=detail_currency,
        hrs_cols=detail_hrs,
        int_cols=detail_int,
        summary_row=True,
        min_widths=detail_widths,
    )

    wb.save(str(output_path))
    logger.info(
        "Master Excel saved: %s (%d claims, %d service lines)",
        output_path, len(summary_df), len(detail_df),
    )
    return output_path


def build_and_write_report(
    parsed_df: pd.DataFrame,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Full pipeline: parsed 835 data → detail DataFrame → claim summary → Excel.

    Returns:
        Path to the generated Excel file.
    """
    # Step 1: Build the detail (SVC-level) DataFrame — source of truth
    detail_df = build_detail_dataframe(parsed_df)

    # Step 2: Derive the claim summary from the detail
    summary_df = build_claim_summary(detail_df)

    # Step 3: Write both to Excel
    return write_master_excel(detail_df, summary_df, output_dir=output_dir)


def rebuild_report_from_cached_csvs(
    processed_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
) -> Optional[Path]:
    """
    Read all parsed CSVs from processed_dir and rebuild the Master Excel report.
    This enables incremental run updates without re-parsing raw .rmt files.
    """
    processed_dir = processed_dir or Config.PROCESSED_DATA_DIR
    csv_files = sorted([f for f in processed_dir.glob("*.csv")])

    if not csv_files:
        logger.warning("No parsed CSVs found in %s to rebuild report.", processed_dir)
        return None

    logger.info("Rebuilding master report from %d cached CSV files...", len(csv_files))

    dfs = []
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file)
            # Ensure proper conversion of NPI and check/EFT columns to string to avoid format truncation issues later
            for col in ["Check/EFT Number", "Member ID", "Rendering Provider NPI", "Claim Control Number (TCN)", "Payee NPI"]:
                if col in df.columns:
                    df[col] = df[col].astype(str)
            dfs.append(df)
        except Exception as e:
            logger.error("Failed to read CSV %s: %s", csv_file.name, e)

    if not dfs:
        logger.warning("No valid CSV data read.")
        return None

    combined_df = pd.concat(dfs, ignore_index=True)
    return build_and_write_report(combined_df, output_dir=output_dir)

